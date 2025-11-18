import asyncio
import httpx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import os  # Import the 'os' library to check if the file exists

API_URL = "https://pfebooks.com/wp-json/wp/v2/posts"
OUTPUT_FILE = "entreprises_pfe.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0"}


async def fetch_page(client, page):
    """Fetches a single page of posts from the API."""
    try:
        r = await client.get(API_URL, params={"per_page": 100, "page": page}, timeout=30)
        r.raise_for_status()
        return r.json()
    except httpx.RequestError as e:
        print(f"An error occurred while requesting page {page}: {e}")
        return []
    except Exception as e:
        print(f"An unexpected error occurred on page {page}: {e}")
        return []

async def scrape_all():
    """Scrapes all post titles and returns them as a list of strings."""
    results = []
    async with httpx.AsyncClient(headers=HEADERS) as client:
        page = 1
        print("Starting scraping...")
        while True:
            data = await fetch_page(client, page)
            if not data:
                break
            for item in data:
                title = item.get("title", {}).get("rendered", "No Title")
                results.append(title)  # Just get the names for now
            print(f"Page {page} scraped successfully.")
    print(f"Scraping finished. Found {len(results)} items.")
    return results


# --- This function has been completely rewritten to be "smart" ---
def save_excel(scraped_names):
    """
    Updates the Excel file by adding only new companies,
    preserving all existing data and modifications.
    """
    
    # --- Step 1: Read the existing Excel file (if it exists) ---
    existing_names = set()
    if os.path.exists(OUTPUT_FILE):
        print(f"Found existing file: {OUTPUT_FILE}. Reading its content.")
        try:
            existing_df = pd.read_excel(OUTPUT_FILE)
            # Store existing names in a set for fast lookups
            existing_names = set(existing_df['Name'].tolist())
        except Exception as e:
            print(f"Could not read the existing Excel file. A new one will be created. Error: {e}")
            existing_df = pd.DataFrame(columns=["Name", "Project Submitted", "Response", "Statut"])
    else:
        print("No existing Excel file found. A new one will be created.")
        existing_df = pd.DataFrame(columns=["Name", "Project Submitted", "Response", "Statut"])

    # --- Step 2: Find only the truly new companies ---
    new_companies_to_add = []
    for name in scraped_names:
        if name not in existing_names:
            new_companies_to_add.append({
                "Name": name,
                "Project Submitted": "",
                "Response": "",
                "Statut": "Non fait"
            })

    if not new_companies_to_add:
        print("No new companies found. The Excel file is already up-to-date.")
        return  # Stop if there's nothing to do

    print(f"Found {len(new_companies_to_add)} new companies to add.")

    # --- Step 3: Append the new companies to the existing data ---
    new_df = pd.DataFrame(new_companies_to_add)
    final_df = pd.concat([existing_df, new_df], ignore_index=True)

    # --- Step 4: Save and format the updated file ---
    final_df.to_excel(OUTPUT_FILE, index=False)
    
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    dv = DataValidation(type="list", formula1='"Non fait,Fait"')
    
    # Apply formatting and validation to a large range to account for future additions
    max_row_buffer = ws.max_row + 500
    dv.add(f"D2:D{max_row_buffer}")
    formatting_range = f"A2:D{max_row_buffer}"
    ws.conditional_formatting.add(formatting_range, FormulaRule(formula=['$D2="Fait"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(formatting_range, FormulaRule(formula=['$D2="Non fait"'], fill=RED_FILL))
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    wb.save(OUTPUT_FILE)
    print(f"✔ Excel file updated successfully: {OUTPUT_FILE}")


async def main():
    """Main function to run the scraper and Excel builder."""
    scraped_names = await scrape_all()
    save_excel(scraped_names)

if __name__ == "__main__":
    asyncio.run(main())
